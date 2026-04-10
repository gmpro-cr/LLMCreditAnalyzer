"""
Multi-pass PDF extraction for financial statements.
3 targeted LLM calls: Balance Sheet | P&L | Cash Flow + Notes
"""
import pdfplumber
import re
import json
import logging
import os
from typing import Dict, Any
import httpx

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

LIMIT = 60000  # ~15-20 pages worth of financial data

# Strong headers — indicate the actual financial statement page (high priority)
BS_STRONG = {
    "balance sheet as at", "balance sheet as on",
    "statement of assets and liabilities", "balance sheet for the year"
}
PL_STRONG = {
    "statement of profit and loss for", "statement of profit and loss\n",
    "profit and loss account for", "profit and loss for the year"
}
CF_STRONG = {
    "cash flow statement for", "statement of cash flow",
    "cash flows from operating activities", "cash flow from operating"
}

# Weak keywords — may appear in TOC, highlights, notes etc.
BS_KEYWORDS = {
    "balance sheet", "total assets", "equity and liabilities",
    "shareholders' funds", "shareholders funds", "inventories",
    "trade receivables", "borrowings", "current assets", "non-current assets"
}
PL_KEYWORDS = {
    "statement of profit", "revenue from operations", "net profit",
    "profit before tax", "finance costs", "depreciation", "total income",
    "employee benefit", "other expenses", "profit after tax"
}
CF_NOTES_KEYWORDS = {
    "cash flow", "notes to", "contingent liabilities", "related party",
    "capital work", "secured loans", "unsecured loans", "working capital"
}


def _score_page(lower: str, strong_kws: set, weak_kws: set) -> int:
    """Score a page: strong header match = 10pts, weak keyword = 1pt each."""
    strong = sum(10 for kw in strong_kws if kw in lower)
    weak = sum(1 for kw in weak_kws if kw in lower)
    return strong + weak


def extract_pages_by_statement(pdf_path: str) -> Dict[str, str]:
    """Score every page for each statement type, then pick highest-scoring pages first."""
    scored: Dict[str, list] = {"balance_sheet": [], "profit_loss": [], "cash_flow_notes": []}
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for i, page in enumerate(pdf.pages):
                text = page.extract_text() or ""
                tables = page.extract_tables()
                table_text = ""
                for table in tables:
                    if table:
                        for row in table:
                            if row:
                                table_text += " | ".join(str(c) if c else "" for c in row) + "\n"
                page_text = text + "\n" + table_text
                lower = page_text.lower()

                bs_score = _score_page(lower, BS_STRONG, BS_KEYWORDS)
                pl_score = _score_page(lower, PL_STRONG, PL_KEYWORDS)
                cf_score = _score_page(lower, CF_STRONG, CF_NOTES_KEYWORDS)

                if bs_score > 0:
                    scored["balance_sheet"].append((bs_score, i, page_text))
                if pl_score > 0:
                    scored["profit_loss"].append((pl_score, i, page_text))
                if cf_score > 0:
                    scored["cash_flow_notes"].append((cf_score, i, page_text))
    except Exception as e:
        logger.error(f"Page classifier error: {e}")

    pages = {"balance_sheet": "", "profit_loss": "", "cash_flow_notes": ""}
    for key, page_list in scored.items():
        # Sort by score descending, then page order for ties
        page_list.sort(key=lambda x: (-x[0], x[1]))
        for score, pg_num, page_text in page_list:
            if len(pages[key]) >= LIMIT:
                break
            pages[key] += f"\n=== PAGE {pg_num + 1} (score={score}) ===\n" + page_text
        pages[key] = pages[key][:LIMIT]
        logger.info(f"  {key}: {len(pages[key])} chars (from {len(page_list)} candidate pages)")
    return pages


def _call_llm(text: str, prompt: str, label: str) -> dict:
    """Send text+prompt to LLM (Gemini primary, Ollama fallback). Returns parsed JSON."""
    provider = os.getenv("EXTRACTION_PROVIDER", "gemini").lower()
    if provider == "gemini":
        result = _call_gemini(text, prompt, label)
        if result:
            return result
        logger.warning(f"[{label}] Gemini failed, falling back to Ollama")
    return _call_ollama(text, prompt, label)


def _call_gemini(text: str, prompt: str, label: str) -> dict:
    try:
        from google import genai
        api_key = os.getenv("GEMINI_API_KEY")
        if not api_key:
            return {}
        client = genai.Client(api_key=api_key)
        full_prompt = prompt + "\n\nTEXT:\n" + text
        response = client.models.generate_content(
            model="models/gemini-flash-lite-latest",
            contents=full_prompt,
        )
        raw = response.text.strip()
        raw = re.sub(r"^```json?\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
        result = json.loads(raw)
        logger.info(f"[{label}] Gemini OK")
        return result
    except Exception as e:
        logger.error(f"[{label}] Gemini error: {e}")
        return {}


def _call_ollama(text: str, prompt: str, label: str) -> dict:
    ollama_url = os.getenv("OLLAMA_URL", "http://localhost:11434")
    ollama_model = os.getenv("OLLAMA_MODEL", "llama3.2:3b")
    try:
        resp = httpx.post(
            f"{ollama_url}/v1/chat/completions",
            json={
                "model": ollama_model,
                "messages": [
                    {"role": "system", "content": "You are a financial data extractor. Return ONLY valid JSON, no markdown, no explanation."},
                    {"role": "user", "content": prompt + "\n\nTEXT:\n" + text},
                ],
                "temperature": 0.1,
                "stream": False,
            },
            timeout=120.0,
        )
        resp.raise_for_status()
        raw = resp.json()["choices"][0]["message"]["content"].strip()
        raw = re.sub(r"^```json?\n?", "", raw)
        raw = re.sub(r"\n?```$", "", raw)
        result = json.loads(raw)
        logger.info(f"[{label}] Ollama OK")
        return result
    except Exception as e:
        logger.error(f"[{label}] Ollama error: {e}")
        return {}


def _extract_balance_sheet(text: str, company_name: str) -> dict:
    prompt = f"""Extract ONLY balance sheet data for {company_name}. Convert all numbers to CRORES.
Return ONLY this JSON (use 0 if not found):
{{
  "balance_sheet": {{
    "current_year": "", "previous_year": "",
    "assets": {{
      "current_assets": {{
        "inventories": {{"raw_materials": 0, "work_in_progress": 0, "finished_goods": 0, "total": 0}},
        "trade_receivables": {{"less_than_6_months": 0, "more_than_6_months": 0, "total": 0}},
        "cash_and_bank": 0, "bank_deposits": 0, "loans_advances": 0,
        "other_current_assets": 0, "total_current_assets": 0
      }},
      "non_current_assets": {{
        "property_plant_equipment": 0, "capital_wip": 0, "intangible_assets": 0,
        "investments": 0, "other_non_current_assets": 0, "total_non_current_assets": 0
      }},
      "total_assets": 0
    }},
    "liabilities": {{
      "current_liabilities": {{
        "short_term_borrowings": 0,
        "trade_payables": {{"msme": 0, "others": 0, "total": 0}},
        "other_current_liabilities": 0, "provisions": 0,
        "current_portion_long_term_debt": 0, "total_current_liabilities": 0
      }},
      "non_current_liabilities": {{
        "long_term_borrowings": 0, "deferred_tax_liability": 0,
        "other_non_current_liabilities": 0, "total_non_current_liabilities": 0
      }},
      "total_liabilities": 0
    }},
    "equity": {{"share_capital": 0, "reserves_surplus": 0, "total_equity": 0}}
  }}
}}"""
    return _call_llm(text, prompt, "balance_sheet")


def _extract_profit_loss(text: str, company_name: str) -> dict:
    prompt = f"""Extract ONLY Profit & Loss data and company info for {company_name}. Convert to CRORES.
Return ONLY this JSON:
{{
  "company_info": {{
    "name": "{company_name}", "cin": "", "financial_year": "",
    "industry": "", "auditor": "", "auditor_opinion": "", "date_of_report": ""
  }},
  "profit_loss": {{
    "revenue": {{
      "revenue_from_operations": 0, "other_income": 0, "total_income": 0
    }},
    "expenses": {{
      "cost_of_materials": 0, "purchases_traded_goods": 0, "change_in_inventory": 0,
      "employee_benefit_expense": 0, "finance_costs": 0,
      "depreciation_amortization": 0, "other_expenses": 0, "total_expenses": 0
    }},
    "profit_metrics": {{
      "ebitda": 0, "ebit": 0, "profit_before_tax": 0,
      "tax_expense": 0, "profit_after_tax": 0
    }}
  }}
}}"""
    return _call_llm(text, prompt, "profit_loss")


def _extract_cash_flow_notes(text: str, company_name: str) -> dict:
    prompt = f"""Extract ONLY cash flow and notes data for {company_name}. Convert to CRORES.
Return ONLY this JSON:
{{
  "cash_flow": {{
    "operating_activities": {{
      "profit_before_tax": 0, "adjustments": 0,
      "working_capital_changes": 0, "taxes_paid": 0, "net_cash_from_operating": 0
    }},
    "investing_activities": {{
      "purchase_of_assets": 0, "sale_of_assets": 0, "net_cash_from_investing": 0
    }},
    "financing_activities": {{
      "proceeds_from_borrowings": 0, "repayment_of_borrowings": 0,
      "interest_paid": 0, "dividends_paid": 0, "net_cash_from_financing": 0
    }},
    "net_change_in_cash": 0, "opening_cash": 0, "closing_cash": 0
  }},
  "notes_analysis": {{
    "contingent_liabilities": {{"bank_guarantees": 0, "legal_cases": 0, "other": 0, "total": 0}},
    "borrowing_details": {{
      "secured_loans": 0, "unsecured_loans": 0,
      "working_capital_limit": 0, "term_loan_outstanding": 0
    }},
    "capex": {{"additions_during_year": 0, "cwip": 0}}
  }},
  "previous_year_data": {{
    "revenue": 0, "net_profit": 0, "total_assets": 0, "total_equity": 0, "total_debt": 0
  }}
}}"""
    return _call_llm(text, prompt, "cash_flow_notes")


def get_empty_structure() -> Dict[str, Any]:
    return {
        "company_info": {
            "name": "", "cin": "", "financial_year": "",
            "industry": "", "auditor": "", "auditor_opinion": "", "date_of_report": ""
        },
        "balance_sheet": {
            "current_year": "", "previous_year": "",
            "assets": {
                "current_assets": {
                    "inventories": {"raw_materials": 0, "work_in_progress": 0, "finished_goods": 0, "total": 0},
                    "trade_receivables": {"less_than_6_months": 0, "more_than_6_months": 0, "total": 0},
                    "cash_and_bank": 0, "bank_deposits": 0, "loans_advances": 0,
                    "other_current_assets": 0, "total_current_assets": 0,
                },
                "non_current_assets": {
                    "property_plant_equipment": 0, "capital_wip": 0, "intangible_assets": 0,
                    "investments": 0, "other_non_current_assets": 0, "total_non_current_assets": 0,
                },
                "total_assets": 0,
            },
            "liabilities": {
                "current_liabilities": {
                    "short_term_borrowings": 0,
                    "trade_payables": {"msme": 0, "others": 0, "total": 0},
                    "other_current_liabilities": 0, "provisions": 0,
                    "current_portion_long_term_debt": 0, "total_current_liabilities": 0,
                },
                "non_current_liabilities": {
                    "long_term_borrowings": 0, "deferred_tax_liability": 0,
                    "other_non_current_liabilities": 0, "total_non_current_liabilities": 0,
                },
                "total_liabilities": 0,
            },
            "equity": {"share_capital": 0, "reserves_surplus": 0, "total_equity": 0},
        },
        "profit_loss": {
            "revenue": {"revenue_from_operations": 0, "other_income": 0, "total_income": 0},
            "expenses": {
                "cost_of_materials": 0, "purchases_traded_goods": 0, "change_in_inventory": 0,
                "employee_benefit_expense": 0, "finance_costs": 0,
                "depreciation_amortization": 0, "other_expenses": 0, "total_expenses": 0,
            },
            "profit_metrics": {
                "ebitda": 0, "ebit": 0, "profit_before_tax": 0,
                "tax_expense": 0, "profit_after_tax": 0,
            },
        },
        "cash_flow": {
            "operating_activities": {
                "profit_before_tax": 0, "adjustments": 0,
                "working_capital_changes": 0, "taxes_paid": 0, "net_cash_from_operating": 0,
            },
            "investing_activities": {"purchase_of_assets": 0, "sale_of_assets": 0, "net_cash_from_investing": 0},
            "financing_activities": {
                "proceeds_from_borrowings": 0, "repayment_of_borrowings": 0,
                "interest_paid": 0, "dividends_paid": 0, "net_cash_from_financing": 0,
            },
            "net_change_in_cash": 0, "opening_cash": 0, "closing_cash": 0,
        },
        "notes_analysis": {
            "contingent_liabilities": {"bank_guarantees": 0, "legal_cases": 0, "other": 0, "total": 0},
            "borrowing_details": {
                "secured_loans": 0, "unsecured_loans": 0,
                "working_capital_limit": 0, "term_loan_outstanding": 0,
            },
            "capex": {"additions_during_year": 0, "cwip": 0},
        },
        "previous_year_data": {"revenue": 0, "net_profit": 0, "total_assets": 0, "total_equity": 0, "total_debt": 0},
    }


def _replace_nulls(data):
    if isinstance(data, dict):
        return {k: _replace_nulls(v) for k, v in data.items()}
    if data is None:
        return 0
    return data


def extract_financials_multi_pass(pdf_path: str, company_name: str = "") -> Dict[str, Any]:
    """3 sequential LLM calls — one per financial statement type. Merges into one dict."""
    logger.info(f"Multi-pass extraction: {pdf_path}")
    pages = extract_pages_by_statement(pdf_path)

    bs_data = _extract_balance_sheet(pages["balance_sheet"], company_name) if pages["balance_sheet"] else {}
    pl_data = _extract_profit_loss(pages["profit_loss"], company_name) if pages["profit_loss"] else {}
    cf_data = _extract_cash_flow_notes(pages["cash_flow_notes"], company_name) if pages["cash_flow_notes"] else {}

    merged = get_empty_structure()
    if pl_data.get("company_info"):
        merged["company_info"] = {**merged["company_info"], **pl_data["company_info"]}
    if bs_data.get("balance_sheet"):
        merged["balance_sheet"] = {**merged["balance_sheet"], **bs_data["balance_sheet"]}
    if pl_data.get("profit_loss"):
        merged["profit_loss"] = {**merged["profit_loss"], **pl_data["profit_loss"]}
    if cf_data.get("cash_flow"):
        merged["cash_flow"] = {**merged["cash_flow"], **cf_data["cash_flow"]}
    if cf_data.get("notes_analysis"):
        merged["notes_analysis"] = {**merged["notes_analysis"], **cf_data["notes_analysis"]}
    if cf_data.get("previous_year_data"):
        merged["previous_year_data"] = {**merged["previous_year_data"], **cf_data["previous_year_data"]}

    # Override company name if provided and not found in PDF
    if company_name and not merged["company_info"].get("name"):
        merged["company_info"]["name"] = company_name

    merged = _replace_nulls(merged)
    logger.info("Multi-pass extraction complete")
    return merged


def extract_excel_cma(file_path: str, company_name: str = "") -> dict:
    """
    Extract financial data from CMA Excel files (.xlsx).
    Looks for standard CMA Data sheets: Balance Sheet, Profit & Loss, Cash Flow.
    Returns same structure as extract_financials_multi_pass for downstream compatibility.
    """
    import openpyxl
    wb = openpyxl.load_workbook(file_path, data_only=True)

    result = {
        "company_info":  {"name": company_name, "source": "excel_cma"},
        "profit_loss":   {},
        "balance_sheet": {},
        "cash_flow":     {},
        "source":        "excel",
    }

    def _cell_val(ws, row, col):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, (int, float)):
            return float(v)
        return None

    def _find_row(ws, keyword, max_rows=120):
        kw = keyword.lower()
        for r in range(1, max_rows):
            for c in range(1, 6):
                v = ws.cell(row=r, column=c).value
                if v and kw in str(v).lower():
                    return r
        return None

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sn = sheet_name.lower()

        if any(k in sn for k in ["p&l", "profit", "income"]):
            rev_row = _find_row(ws, "revenue from operations") or _find_row(ws, "net sales")
            pat_row = _find_row(ws, "profit after tax") or _find_row(ws, "net profit")
            ebt_row = _find_row(ws, "ebitda") or _find_row(ws, "operating profit")
            if rev_row:
                result["profit_loss"]["revenue"] = {
                    "revenue_from_operations": {"current": _cell_val(ws, rev_row, 3)}
                }
            if pat_row:
                result["profit_loss"].setdefault("profit_metrics", {})["profit_after_tax"] = \
                    {"current": _cell_val(ws, pat_row, 3)}
            if ebt_row:
                result["profit_loss"].setdefault("profit_metrics", {})["ebitda"] = \
                    {"current": _cell_val(ws, ebt_row, 3)}

        elif any(k in sn for k in ["balance", "b/s", "bs"]):
            ta_row  = _find_row(ws, "total assets")
            eq_row  = _find_row(ws, "net worth") or _find_row(ws, "shareholders equity")
            dbt_row = _find_row(ws, "total borrowings") or _find_row(ws, "term loan")
            ca_row  = _find_row(ws, "total current assets") or _find_row(ws, "current assets")
            cl_row  = _find_row(ws, "total current liabilities") or _find_row(ws, "current liabilities")
            if ta_row:
                result["balance_sheet"]["assets"] = {
                    "total_assets": {"current": _cell_val(ws, ta_row, 3)}
                }
            if eq_row:
                result["balance_sheet"]["equity"] = {"total_equity": _cell_val(ws, eq_row, 3)}
            if dbt_row:
                result["balance_sheet"]["liabilities"] = {
                    "borrowings": {"total_borrowings": {"current": _cell_val(ws, dbt_row, 3)}}
                }
            if ca_row:
                result["balance_sheet"].setdefault("current_assets", {})["total_current_assets"] = \
                    {"current": _cell_val(ws, ca_row, 3)}
            if cl_row:
                result["balance_sheet"].setdefault("current_liabilities", {})["total_current_liabilities"] = \
                    {"current": _cell_val(ws, cl_row, 3)}

    return result

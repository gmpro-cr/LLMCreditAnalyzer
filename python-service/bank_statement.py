"""
Bank statement analyzer — banker-grade analysis aligned with Indian credit underwriting practice.

Outputs:
  - Account & period summary
  - Inflow/outflow totals, monthly trend
  - Average / minimum / maximum balance, AMB, balance volatility
  - Recurring credits (salary / business inflow) and recurring debits (EMI / SI)
  - Cash deposits & withdrawals
  - Cheque return / bounce detection
  - Negative balance & OD utilization days
  - Counterparty concentration
  - Risk flags & weighted creditworthiness score (0-100)

Accepts CSV, Excel, or PDF (text-based).
"""
from __future__ import annotations

import io
import re
import logging
import statistics
from dataclasses import dataclass
from datetime import datetime, date
from collections import defaultdict, Counter
from typing import Optional, Any

import pandas as pd

logger = logging.getLogger(__name__)

# ── Heuristics ────────────────────────────────────────────────────────────────

CASH_KEYWORDS = re.compile(r"\b(cash|cdm|atm-?(?:dep|wdl)|cash\s*deposit|cash\s*withdrawal)\b", re.I)
SALARY_KEYWORDS = re.compile(r"\b(salary|sal[\s/-]+|payroll|stipend|wages)\b", re.I)
EMI_KEYWORDS = re.compile(r"\b(emi|loan|nbfc|bajaj\s*fin|hdfc\s*loan|tata\s*cap|hdb|chola|muthoot|equitas|idfc\s*loan|home\s*loan|car\s*loan|personal\s*loan|aditya\s*birla\s*fin)\b", re.I)
BOUNCE_KEYWORDS = re.compile(r"\b(chq?\s*ret|chq?\s*rtn|return(?:ed)?\s*chq|return\s*charge|insufficient\s*funds|i/?w\s*ret|o/?w\s*ret|ecs\s*return|ach\s*return|nach\s*return|inward\s*return|cheque\s*bounce|bounce\s*charge|chargeback|cheque\s*dishonou?r)\b", re.I)
OD_KEYWORDS = re.compile(r"\b(overdraft|od\s*int|cash\s*credit|cc\s*int)\b", re.I)
GST_KEYWORDS = re.compile(r"\b(gst|tax\s*paid|tds)\b", re.I)
INVESTMENT_KEYWORDS = re.compile(r"\b(mutual\s*fund|mf\s*sip|sip|nps|ppf|fd\s*close|fd\s*open|rd\s*credit)\b", re.I)
UPI_KEYWORDS = re.compile(r"\b(upi|gpay|phonepe|paytm|bhim|imps|neft|rtgs)\b", re.I)


@dataclass
class Transaction:
    txn_date: date
    description: str
    amount: float          # signed: positive = credit, negative = debit
    balance: Optional[float]
    raw: str

    @property
    def is_credit(self) -> bool:
        return self.amount > 0

    @property
    def is_debit(self) -> bool:
        return self.amount < 0

    @property
    def magnitude(self) -> float:
        return abs(self.amount)


# ── Parsers ───────────────────────────────────────────────────────────────────

def _parse_date(s: Any) -> Optional[date]:
    if s is None:
        return None
    if isinstance(s, (datetime, pd.Timestamp)):
        return s.date()
    if isinstance(s, date):
        return s
    s = str(s).strip()
    for fmt in ("%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y", "%Y-%m-%d", "%d %b %Y", "%d-%b-%Y", "%d %b %y", "%d.%m.%Y", "%d.%m.%y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(v: Any) -> Optional[float]:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    s = str(v).strip().replace(",", "").replace("₹", "").replace("INR", "").replace("Rs.", "").replace("Rs", "")
    if not s or s.lower() in ("nan", "-", "—"):
        return None
    s = s.replace("(", "-").replace(")", "")
    try:
        return float(s)
    except ValueError:
        return None


def _detect_columns(cols: list[str]) -> dict[str, Optional[str]]:
    """Map header names to our standard fields."""
    norm = {c: c.lower().strip().replace("_", " ").replace("-", " ") for c in cols}
    out: dict[str, Optional[str]] = {"date": None, "desc": None, "debit": None, "credit": None, "amount": None, "balance": None}

    for c, n in norm.items():
        if out["date"] is None and re.search(r"\b(txn|transaction|value|posting|tran)\s*date|^date$|^dt$", n):
            out["date"] = c
        elif out["date"] is None and "date" in n:
            out["date"] = c
        if out["desc"] is None and re.search(r"description|narration|particular|details|remarks|narrate", n):
            out["desc"] = c
        if out["debit"] is None and re.search(r"debit|withdraw|dr\b|paid out", n):
            out["debit"] = c
        if out["credit"] is None and re.search(r"credit|deposit|cr\b|paid in", n):
            out["credit"] = c
        if out["amount"] is None and n in ("amount", "amt", "txn amount", "transaction amount"):
            out["amount"] = c
        if out["balance"] is None and re.search(r"balance|closing", n):
            out["balance"] = c
    return out


def _parse_dataframe(df: pd.DataFrame) -> list[Transaction]:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    cols = _detect_columns(list(df.columns))
    if not cols["date"] or not cols["desc"]:
        # try heuristic: assume first column is date, second is description
        if len(df.columns) >= 3:
            cols["date"] = df.columns[0]
            cols["desc"] = df.columns[1]

    txns: list[Transaction] = []
    for _, row in df.iterrows():
        d = _parse_date(row.get(cols["date"])) if cols["date"] else None
        if d is None:
            continue
        desc = str(row.get(cols["desc"], "") or "").strip() if cols["desc"] else ""
        if not desc and cols["desc"] is None:
            desc = " ".join(str(v) for v in row.values if v is not None)

        debit = _to_float(row.get(cols["debit"])) if cols["debit"] else None
        credit = _to_float(row.get(cols["credit"])) if cols["credit"] else None
        amount: Optional[float] = None
        if debit and debit > 0:
            amount = -abs(debit)
        elif credit and credit > 0:
            amount = abs(credit)
        elif cols["amount"]:
            v = _to_float(row.get(cols["amount"]))
            if v is not None:
                # If we have a separate Dr/Cr indicator column we'd parse it; default: positive=credit
                amount = v
        if amount is None or amount == 0:
            continue

        bal = _to_float(row.get(cols["balance"])) if cols["balance"] else None
        txns.append(Transaction(d, desc, amount, bal, raw=" | ".join(str(v) for v in row.values)))
    return txns


# ── PDF parser (text-based, line heuristics) ──────────────────────────────────

PDF_LINE_RE = re.compile(
    r"^(?P<date>\d{1,2}[/-][A-Za-z0-9]{1,4}[/-]\d{2,4})\s+"
    r"(?P<desc>.+?)\s+"
    r"(?:(?P<debit>[\d,]+\.\d{2})\s+(?P<credit>[\d,]+\.\d{2})|(?P<amount>-?[\d,]+\.\d{2}))"
    r"\s+(?P<balance>[\d,]+\.\d{2})\s*(?:Cr|Dr)?\s*$",
    re.I,
)

DATE_RE = re.compile(r"^\d{1,2}[./-]\d{1,2}[./-]\d{2,4}$")
AMOUNT_RE = re.compile(r"^-?[\d,]+\.\d{2}$")


def _parse_pdf_positional(pdf_bytes: bytes) -> list[Transaction]:
    """
    Position-aware PDF parser for bank statements (e.g. ICICI) where pdfplumber
    extracts words with x/y coordinates. Detects column boundaries from the
    header row and classifies each word accordingly.
    """
    import pdfplumber

    txns: list[Transaction] = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            words = page.extract_words(x_tolerance=4, y_tolerance=4)
            if not words:
                continue

            # ── 1. Detect column boundaries from header keywords ──────────────
            # Look for header words: "Withdrawal", "Deposit", "Balance"
            withdrawal_x: Optional[float] = None
            deposit_x: Optional[float] = None
            balance_x: Optional[float] = None
            date_x: Optional[float] = None

            for w in words:
                txt = w["text"].lower()
                if txt in ("withdrawal", "debit", "dr") and withdrawal_x is None:
                    withdrawal_x = w["x0"]
                elif txt in ("deposit", "credit", "cr") and deposit_x is None:
                    deposit_x = w["x0"]
                elif txt == "balance" and balance_x is None:
                    balance_x = w["x0"]
                elif txt == "date" and date_x is None:
                    date_x = w["x0"]

            # Fallback boundaries if header not found
            if withdrawal_x is None:
                withdrawal_x = 390.0
            if deposit_x is None:
                deposit_x = 460.0
            if balance_x is None:
                balance_x = 530.0

            # ── 2. Find transaction rows anchored by date words ───────────────
            # Group all words by rounded y (±4px tolerance)
            rows: dict[int, list[dict]] = {}
            for w in words:
                key = round(w["top"] / 5) * 5
                rows.setdefault(key, []).append(w)

            sorted_ys = sorted(rows.keys())

            # Collect date-row anchors
            date_rows: list[tuple[int, dict]] = []
            for y_key in sorted_ys:
                for w in rows[y_key]:
                    if DATE_RE.match(w["text"]):
                        date_rows.append((y_key, w))
                        break

            # ── 3. For each date row, collect amounts + nearby description ────
            for i, (y_key, date_word) in enumerate(date_rows):
                d = _parse_date(date_word["text"])
                if not d:
                    continue

                row_words = rows[y_key]

                # Split words into columns by x position
                withdrawal_amt: Optional[float] = None
                deposit_amt: Optional[float] = None
                balance_val: Optional[float] = None

                for w in row_words:
                    if not AMOUNT_RE.match(w["text"]):
                        continue
                    v = _to_float(w["text"])
                    if v is None:
                        continue
                    x = w["x0"]
                    if x >= balance_x - 20:
                        balance_val = v
                    elif x >= deposit_x - 20:
                        deposit_amt = v
                    elif x >= withdrawal_x - 20:
                        withdrawal_amt = v

                # Determine signed amount
                amount: Optional[float] = None
                if withdrawal_amt and withdrawal_amt > 0:
                    amount = -abs(withdrawal_amt)
                elif deposit_amt and deposit_amt > 0:
                    amount = abs(deposit_amt)
                elif balance_val is not None and txns:
                    # Infer from balance delta
                    prev_bal = txns[-1].balance
                    if prev_bal is not None:
                        amount = round(balance_val - prev_bal, 2)

                if amount is None or amount == 0:
                    continue

                # Collect description using a fixed window around the date row.
                # ICICI layout: narration line appears just ABOVE the S.No/date line,
                # with 0-2 continuation lines just BELOW. Row pitch is ~30px.
                # We look ±(row_pitch - 5)px so we never bleed into adjacent transactions.
                row_pitch = 30  # typical px between consecutive ICICI rows
                if len(date_rows) >= 2:
                    pitches = [date_rows[k + 1][0] - date_rows[k][0]
                               for k in range(min(5, len(date_rows) - 1))]
                    row_pitch = int(statistics.median(pitches))
                window = max(row_pitch - 5, 10)

                desc_parts: list[str] = []

                def _is_desc_word(w: dict) -> bool:
                    return (w["x0"] > 100 and w["x0"] < withdrawal_x - 5
                            and not DATE_RE.match(w["text"])
                            and not w["text"].isdigit()
                            and not AMOUNT_RE.match(w["text"]))

                # Narration appears at diff=-5 (above date row) and diff=+5,+15 (below).
                # diff=-25/-15 belongs to previous txn; diff=+25 belongs to next txn.
                # Use [-8, +18] window to capture exactly this transaction's narration.
                for y2 in sorted_ys:
                    diff = y2 - y_key
                    if diff < -8 or diff > 18:
                        continue
                    if y2 == y_key:
                        continue
                    desc_parts.extend(w["text"] for w in rows[y2] if _is_desc_word(w))

                desc = " ".join(desc_parts).strip()
                if not desc:
                    desc = date_word["text"]

                txns.append(Transaction(d, desc, amount, balance_val, raw=desc))

    return txns


def _parse_pdf_text(text: str) -> list[Transaction]:
    txns: list[Transaction] = []
    last_balance: Optional[float] = None
    for raw_line in text.splitlines():
        line = raw_line.strip()
        if not line:
            continue
        m = PDF_LINE_RE.match(line)
        if not m:
            continue
        d = _parse_date(m.group("date"))
        if not d:
            continue
        desc = m.group("desc").strip()
        balance = _to_float(m.group("balance"))
        debit = _to_float(m.group("debit"))
        credit = _to_float(m.group("credit"))
        amount: Optional[float] = None
        if debit and debit > 0:
            amount = -abs(debit)
        elif credit and credit > 0:
            amount = abs(credit)
        elif m.group("amount"):
            v = _to_float(m.group("amount"))
            if v is not None:
                # disambiguate via balance delta if available
                if last_balance is not None and balance is not None:
                    amount = balance - last_balance
                else:
                    amount = v
        if amount is None or amount == 0:
            continue
        last_balance = balance if balance is not None else last_balance
        txns.append(Transaction(d, desc, amount, balance, raw=line))
    return txns


def parse_statement(file_bytes: bytes, filename: str) -> list[Transaction]:
    name = filename.lower()
    if name.endswith(".csv"):
        df = pd.read_csv(io.BytesIO(file_bytes), thousands=",", on_bad_lines="skip")
        return _parse_dataframe(df)
    if name.endswith(".xlsx") or name.endswith(".xls"):
        # try header detection — many bank exports have garbage rows at top
        for header_row in (0, 1, 2, 3, 5, 8, 10, 15, 20):
            try:
                df = pd.read_excel(io.BytesIO(file_bytes), header=header_row)
                # require at least one date-like column
                cols = _detect_columns(list(df.columns))
                if cols["date"] and cols["desc"]:
                    txns = _parse_dataframe(df)
                    if len(txns) >= 3:
                        return txns
            except Exception:
                continue
        df = pd.read_excel(io.BytesIO(file_bytes))
        return _parse_dataframe(df)
    if name.endswith(".pdf"):
        try:
            # Try position-aware parser first (handles ICICI and similar multi-column layouts)
            txns = _parse_pdf_positional(file_bytes)
            if len(txns) >= 2:
                logger.info(f"Positional PDF parser extracted {len(txns)} transactions")
                return txns
        except Exception as e:
            logger.warning(f"Positional PDF parser failed: {e}")

        try:
            # Fallback: text-line regex parser
            import pdfplumber
            text_chunks: list[str] = []
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                for page in pdf.pages:
                    t = page.extract_text() or ""
                    text_chunks.append(t)
            return _parse_pdf_text("\n".join(text_chunks))
        except Exception as e:
            logger.warning(f"pdfplumber text parser failed: {e}")
            return []
    raise ValueError(f"Unsupported file type: {filename}")


# ── Analyzers ─────────────────────────────────────────────────────────────────

def _filter_period(txns: list[Transaction], from_date: Optional[date], to_date: Optional[date]) -> list[Transaction]:
    if from_date is None and to_date is None:
        return txns
    out = []
    for t in txns:
        if from_date and t.txn_date < from_date:
            continue
        if to_date and t.txn_date > to_date:
            continue
        out.append(t)
    return out


def _month_key(d: date) -> str:
    return d.strftime("%Y-%m")


def _round(v: Optional[float], digits: int = 2) -> Optional[float]:
    if v is None:
        return None
    return round(v, digits)


def _detect_recurring(txns: list[Transaction], sign: str, min_occurrences: int = 3) -> list[dict]:
    """Group txns by canonical-counterparty key; flag those repeating ≥ min_occurrences with stable amount."""
    groups: dict[str, list[Transaction]] = defaultdict(list)
    for t in txns:
        if sign == "credit" and not t.is_credit:
            continue
        if sign == "debit" and not t.is_debit:
            continue
        key = _canonical_party(t.description)
        if not key:
            continue
        groups[key].append(t)

    recurring = []
    for key, items in groups.items():
        if len(items) < min_occurrences:
            continue
        amounts = [t.magnitude for t in items]
        median = statistics.median(amounts)
        # require variance < 30% of median for "recurring"
        if median > 0 and (statistics.pstdev(amounts) / median) > 0.5:
            continue
        recurring.append({
            "counterparty": key,
            "occurrences": len(items),
            "median_amount": _round(median),
            "total_amount": _round(sum(amounts)),
            "first_seen": min(t.txn_date for t in items).isoformat(),
            "last_seen": max(t.txn_date for t in items).isoformat(),
            "sample_description": items[0].description[:120],
        })
    return sorted(recurring, key=lambda r: -r["total_amount"])


def _canonical_party(desc: str) -> str:
    s = desc.upper()
    s = re.sub(r"\b\d{6,}\b", "", s)         # strip long numbers (txn ids)
    s = re.sub(r"[^A-Z0-9 /&-]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    # take first 4 words as a stable signature
    words = [w for w in s.split() if len(w) > 1][:4]
    return " ".join(words)


def _bounces(txns: list[Transaction]) -> list[dict]:
    out = []
    for t in txns:
        if BOUNCE_KEYWORDS.search(t.description):
            out.append({
                "date": t.txn_date.isoformat(),
                "description": t.description[:160],
                "amount": _round(t.magnitude),
                "type": "outward" if t.is_debit else "inward",
            })
    return out


def _classify_txn(t: Transaction) -> dict:
    """Banker-style category + channel labels for one transaction."""
    desc = t.description
    is_credit = t.is_credit
    category = "Other"
    if BOUNCE_KEYWORDS.search(desc):
        category = "Cheque/ECS Return"
    elif CASH_KEYWORDS.search(desc):
        category = "Cash Deposit" if is_credit else "Cash Withdrawal"
    elif EMI_KEYWORDS.search(desc) and not is_credit:
        category = "EMI / Loan"
    elif SALARY_KEYWORDS.search(desc) and is_credit:
        category = "Salary / Primary Income"
    elif OD_KEYWORDS.search(desc):
        category = "OD / CC Interest"
    elif GST_KEYWORDS.search(desc):
        category = "Tax / GST"
    elif INVESTMENT_KEYWORDS.search(desc):
        category = "Investment"
    elif UPI_KEYWORDS.search(desc):
        category = "Digital Transfer (UPI/IMPS/NEFT/RTGS)"
    elif is_credit:
        category = "Other Credit"
    else:
        category = "Other Debit"

    if re.search(r"\bUPI\b|GPAY|PHONEPE|PAYTM|BHIM", desc, re.I):
        channel = "UPI"
    elif re.search(r"\bIMPS\b", desc, re.I):
        channel = "IMPS"
    elif re.search(r"\bNEFT\b", desc, re.I):
        channel = "NEFT"
    elif re.search(r"\bRTGS\b", desc, re.I):
        channel = "RTGS"
    elif re.search(r"\bATM\b|CDM\b", desc, re.I):
        channel = "ATM/CDM"
    elif re.search(r"\bCHQ\b|CHEQUE", desc, re.I):
        channel = "Cheque"
    elif re.search(r"\bECS\b|ACH\b|NACH\b|MANDATE", desc, re.I):
        channel = "ECS/NACH"
    elif re.search(r"\bSI\b|STANDING\s*INSTR", desc, re.I):
        channel = "Standing Instruction"
    else:
        channel = "Other"

    return {"category": category, "channel": channel, "counterparty": _canonical_party(desc)}


def _cash_txns(txns: list[Transaction]) -> dict:
    cash_in = [t for t in txns if t.is_credit and CASH_KEYWORDS.search(t.description)]
    cash_out = [t for t in txns if t.is_debit and CASH_KEYWORDS.search(t.description)]
    return {
        "cash_deposits": {
            "count": len(cash_in),
            "total": _round(sum(t.magnitude for t in cash_in)),
            "max_single": _round(max((t.magnitude for t in cash_in), default=0)),
        },
        "cash_withdrawals": {
            "count": len(cash_out),
            "total": _round(sum(t.magnitude for t in cash_out)),
            "max_single": _round(max((t.magnitude for t in cash_out), default=0)),
        },
    }


def _emi_obligations(txns: list[Transaction]) -> list[dict]:
    debits = [t for t in txns if t.is_debit and EMI_KEYWORDS.search(t.description)]
    return _detect_recurring(debits, sign="debit", min_occurrences=2)


def _balance_metrics(txns: list[Transaction]) -> dict:
    balances = [t.balance for t in txns if t.balance is not None]
    if not balances:
        return {"average": None, "minimum": None, "maximum": None, "amb": None, "negative_days": 0, "volatility": None}

    # AMB: average of month-end (last txn balance per month)
    by_month: dict[str, list[Transaction]] = defaultdict(list)
    for t in txns:
        if t.balance is not None:
            by_month[_month_key(t.txn_date)].append(t)
    month_end = []
    for _, items in by_month.items():
        items.sort(key=lambda x: x.txn_date)
        month_end.append(items[-1].balance)
    amb = statistics.mean(month_end) if month_end else None

    return {
        "average": _round(statistics.mean(balances)),
        "minimum": _round(min(balances)),
        "maximum": _round(max(balances)),
        "amb": _round(amb),
        "volatility_pct": _round((statistics.pstdev(balances) / abs(statistics.mean(balances)) * 100) if statistics.mean(balances) else None),
        "negative_days": sum(1 for b in balances if b < 0),
    }


def _monthly_trend(txns: list[Transaction]) -> list[dict]:
    rows: dict[str, dict] = defaultdict(lambda: {"inflow": 0.0, "outflow": 0.0, "txn_count": 0, "closing_balance": None})
    last_balance_by_month: dict[str, tuple[date, float]] = {}
    for t in txns:
        m = _month_key(t.txn_date)
        if t.is_credit:
            rows[m]["inflow"] += t.magnitude
        else:
            rows[m]["outflow"] += t.magnitude
        rows[m]["txn_count"] += 1
        if t.balance is not None:
            prior = last_balance_by_month.get(m)
            if prior is None or t.txn_date >= prior[0]:
                last_balance_by_month[m] = (t.txn_date, t.balance)

    out = []
    for m in sorted(rows.keys()):
        cb = last_balance_by_month.get(m)
        out.append({
            "month": m,
            "inflow": _round(rows[m]["inflow"]),
            "outflow": _round(rows[m]["outflow"]),
            "net": _round(rows[m]["inflow"] - rows[m]["outflow"]),
            "txn_count": rows[m]["txn_count"],
            "closing_balance": _round(cb[1]) if cb else None,
        })
    return out


def _counterparty_concentration(txns: list[Transaction], top_n: int = 5) -> dict:
    inflow: dict[str, float] = defaultdict(float)
    outflow: dict[str, float] = defaultdict(float)
    for t in txns:
        key = _canonical_party(t.description)
        if not key:
            continue
        if t.is_credit:
            inflow[key] += t.magnitude
        else:
            outflow[key] += t.magnitude

    def _top(d: dict[str, float]) -> list[dict]:
        items = sorted(d.items(), key=lambda kv: -kv[1])[:top_n]
        total = sum(d.values()) or 1
        return [{"counterparty": k, "amount": _round(v), "share_pct": _round(v / total * 100, 1)} for k, v in items]

    return {"top_inflow_counterparties": _top(inflow), "top_outflow_counterparties": _top(outflow)}


def _risk_flags_and_score(summary: dict) -> tuple[list[dict], int]:
    """
    Bank-credit-team scoring (0-100, higher = healthier).
    Hard cuts: bounces > 3, negative balance > 5 days, cash dependency > 50% of credits.
    """
    flags: list[dict] = []
    score = 100

    bounces = summary.get("cheque_returns", {}).get("count", 0)
    if bounces > 0:
        sev = "high" if bounces >= 3 else "medium"
        score -= 25 if bounces >= 3 else 10
        flags.append({
            "code": "CHEQUE_BOUNCES",
            "severity": sev,
            "title": f"{bounces} cheque/ECS return(s) detected",
            "detail": "Bounce history materially impairs credit eligibility; underwriters typically reject if ≥3 bounces in 12 months.",
        })

    inflow = summary["totals"]["total_inflow"] or 0
    outflow = summary["totals"]["total_outflow"] or 0
    cash_in = summary.get("cash_activity", {}).get("cash_deposits", {}).get("total") or 0
    cash_dep_share = (cash_in / inflow * 100) if inflow else 0
    if cash_dep_share > 50:
        score -= 20
        flags.append({
            "code": "HIGH_CASH_DEPENDENCY",
            "severity": "high",
            "title": f"{cash_dep_share:.0f}% of credits are cash deposits",
            "detail": "High cash inflow dependence is treated as low-quality income — banks heavily discount it for repayment capacity.",
        })
    elif cash_dep_share > 25:
        score -= 8
        flags.append({
            "code": "MODERATE_CASH_DEPENDENCY",
            "severity": "medium",
            "title": f"{cash_dep_share:.0f}% of credits are cash deposits",
            "detail": "Watch — moderate cash dependency reduces verifiable income for underwriting.",
        })

    neg_days = summary["balance_metrics"].get("negative_days") or 0
    if neg_days > 5:
        score -= 15
        flags.append({
            "code": "NEGATIVE_BALANCE",
            "severity": "high",
            "title": f"Account stayed negative for {neg_days} day(s)",
            "detail": "Sustained negative balance indicates liquidity stress and potential OD over-utilization.",
        })
    elif neg_days > 0:
        score -= 5
        flags.append({
            "code": "NEGATIVE_BALANCE_OCCASIONAL",
            "severity": "low",
            "title": f"Occasional negative balance ({neg_days} day(s))",
            "detail": "Brief negative-balance episodes — verify causes with the borrower.",
        })

    amb = summary["balance_metrics"].get("amb") or 0
    if amb > 0 and outflow > 0:
        amb_to_outflow = (amb / (outflow / 12 if len(summary["monthly_trend"]) else outflow)) if summary["monthly_trend"] else 0
        if amb_to_outflow < 0.25:
            score -= 10
            flags.append({
                "code": "LOW_AMB",
                "severity": "medium",
                "title": "Average Monthly Balance is low vs monthly outflow",
                "detail": "AMB < 25% of average monthly outflow suggests thin liquidity buffer.",
            })

    emi_total = sum(e["total_amount"] for e in summary.get("emi_obligations", []))
    if inflow > 0:
        leverage = emi_total / inflow * 100
        if leverage > 50:
            score -= 15
            flags.append({
                "code": "HIGH_LEVERAGE",
                "severity": "high",
                "title": f"EMI obligations are {leverage:.0f}% of total inflow",
                "detail": "Existing EMI burden > 50% of credits — adding new debt likely breaches FOIR norms (typical limit 50–60%).",
            })
        elif leverage > 35:
            score -= 7
            flags.append({
                "code": "MODERATE_LEVERAGE",
                "severity": "medium",
                "title": f"EMI obligations are {leverage:.0f}% of total inflow",
                "detail": "Moderate existing leverage — model FOIR carefully before sanctioning fresh exposure.",
            })

    if not summary.get("recurring_credits"):
        score -= 8
        flags.append({
            "code": "NO_RECURRING_INCOME",
            "severity": "medium",
            "title": "No clear recurring income pattern detected",
            "detail": "No salary or recurring business inflow identified — income source unverifiable from this statement alone.",
        })

    months = len(summary["monthly_trend"])
    if months > 1:
        inflows = [m["inflow"] for m in summary["monthly_trend"] if m["inflow"]]
        if inflows and statistics.mean(inflows):
            cv = statistics.pstdev(inflows) / statistics.mean(inflows)
            if cv > 0.6:
                score -= 8
                flags.append({
                    "code": "VOLATILE_INFLOW",
                    "severity": "medium",
                    "title": "Inflow volatility is high (CV > 60%)",
                    "detail": "Large month-to-month inflow swings — income stability is weak; treat with caution.",
                })

    return flags, max(0, min(100, score))


def _classify(score: int) -> str:
    if score >= 80:
        return "Strong"
    if score >= 65:
        return "Acceptable"
    if score >= 50:
        return "Borderline"
    return "High Risk"


# ── Public entry point ────────────────────────────────────────────────────────

def analyze_bank_statement(
    file_bytes: bytes,
    filename: str,
    period_from: Optional[str] = None,
    period_to: Optional[str] = None,
    account_holder: Optional[str] = None,
) -> dict:
    """
    Analyze a bank statement file and return a structured banker-grade report.
    period_from / period_to are ISO-8601 dates (YYYY-MM-DD).
    """
    txns = parse_statement(file_bytes, filename)
    if not txns:
        raise ValueError("No transactions could be parsed from the file. Ensure it's a text-based PDF/CSV/Excel with date, description, amount and balance columns.")

    from_d = _parse_date(period_from) if period_from else None
    to_d = _parse_date(period_to) if period_to else None
    txns = _filter_period(txns, from_d, to_d)
    if not txns:
        raise ValueError("No transactions fall within the selected period.")

    txns.sort(key=lambda t: t.txn_date)
    inflow_total = sum(t.magnitude for t in txns if t.is_credit)
    outflow_total = sum(t.magnitude for t in txns if t.is_debit)

    bounces = _bounces(txns)
    cash = _cash_txns(txns)
    balance_metrics = _balance_metrics(txns)
    monthly = _monthly_trend(txns)
    recurring_credits = _detect_recurring([t for t in txns if not CASH_KEYWORDS.search(t.description)], sign="credit", min_occurrences=2)
    recurring_debits = _detect_recurring([t for t in txns if not CASH_KEYWORDS.search(t.description) and not BOUNCE_KEYWORDS.search(t.description)], sign="debit", min_occurrences=3)
    emi = _emi_obligations(txns)
    salary_credits = [r for r in recurring_credits if SALARY_KEYWORDS.search(r["sample_description"])]
    concentration = _counterparty_concentration(txns)

    classified_ledger = []
    for t in txns:
        cls = _classify_txn(t)
        classified_ledger.append({
            "date": t.txn_date.isoformat(),
            "description": t.description,
            "category": cls["category"],
            "channel": cls["channel"],
            "counterparty": cls["counterparty"],
            "type": "Credit" if t.is_credit else "Debit",
            "credit": _round(t.magnitude) if t.is_credit else None,
            "debit": _round(t.magnitude) if t.is_debit else None,
            "amount_signed": _round(t.amount),
            "running_balance": _round(t.balance),
        })

    summary = {
        "account": {
            "holder": account_holder or "—",
            "filename": filename,
            "period": {
                "from": (from_d or txns[0].txn_date).isoformat(),
                "to": (to_d or txns[-1].txn_date).isoformat(),
                "months": len(monthly),
            },
            "transactions": len(txns),
        },
        "totals": {
            "total_inflow": _round(inflow_total),
            "total_outflow": _round(outflow_total),
            "net_cash_flow": _round(inflow_total - outflow_total),
            "credit_count": sum(1 for t in txns if t.is_credit),
            "debit_count": sum(1 for t in txns if t.is_debit),
            "opening_balance": _round(txns[0].balance) if txns[0].balance is not None else None,
            "closing_balance": _round(txns[-1].balance) if txns[-1].balance is not None else None,
        },
        "balance_metrics": balance_metrics,
        "monthly_trend": monthly,
        "salary_or_primary_income": salary_credits or recurring_credits[:3],
        "recurring_credits": recurring_credits,
        "recurring_debits": recurring_debits,
        "emi_obligations": emi,
        "cash_activity": cash,
        "cheque_returns": {
            "count": len(bounces),
            "items": bounces[:20],
        },
        "counterparty_concentration": concentration,
        "transactions": classified_ledger,
    }

    flags, score = _risk_flags_and_score(summary)
    summary["risk_flags"] = flags
    summary["score"] = {
        "value": score,
        "rating": _classify(score),
        "scale": "0-100 (higher is healthier; combines bounces, cash dependency, leverage, AMB, income stability)",
    }

    # Underwriter narrative
    summary["underwriter_observations"] = _build_observations(summary)
    return summary


def _build_observations(s: dict) -> list[str]:
    obs: list[str] = []
    a = s["account"]
    t = s["totals"]
    bm = s["balance_metrics"]
    obs.append(
        f"Reviewed {a['transactions']} transactions over {a['period']['months']} month(s) "
        f"({a['period']['from']} to {a['period']['to']})."
    )
    if t["total_inflow"]:
        obs.append(
            f"Total credits ₹{t['total_inflow']:,.0f}; total debits ₹{t['total_outflow']:,.0f}; "
            f"net cash flow ₹{t['net_cash_flow']:,.0f}."
        )
    if bm.get("amb") is not None:
        obs.append(
            f"Average Monthly Balance (AMB) ₹{bm['amb']:,.0f}; min ₹{bm['minimum']:,.0f}, max ₹{bm['maximum']:,.0f}."
        )
    if s["salary_or_primary_income"]:
        top = s["salary_or_primary_income"][0]
        obs.append(
            f"Primary recurring inflow appears to be '{top['counterparty']}' "
            f"(median ₹{top['median_amount']:,.0f} × {top['occurrences']} times)."
        )
    if s["emi_obligations"]:
        total_emi = sum(e["total_amount"] for e in s["emi_obligations"])
        obs.append(f"Existing EMI/loan outflows total ₹{total_emi:,.0f} across {len(s['emi_obligations'])} counterparties.")
    if s["cheque_returns"]["count"]:
        obs.append(f"⚠ {s['cheque_returns']['count']} cheque/ECS return(s) detected — investigate before approval.")
    obs.append(f"Creditworthiness score: {s['score']['value']}/100 ({s['score']['rating']}).")
    return obs


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(analysis: dict) -> bytes:
    """
    Build a single multi-sheet workbook containing:
      - Summary           (account, period, score, totals, balance metrics)
      - Risk Flags
      - Underwriter Notes
      - Monthly Trend
      - Transactions      (full classified ledger)
      - Recurring Credits / Recurring Debits / EMI Obligations
      - Cash Activity
      - Cheque Returns
      - Counterparty Concentration
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    HEADER_FILL = PatternFill("solid", fgColor="0F172A")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    SUB_FILL = PatternFill("solid", fgColor="E2E8F0")
    SUB_FONT = Font(bold=True, color="0F172A", size=11)
    SECTION_FONT = Font(bold=True, size=13, color="0F172A")
    THIN = Side(border_style="thin", color="CBD5E1")
    BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    INR_FMT = '_-₹* #,##,##0_-;_-₹* -#,##,##0_-;_-₹* "-"_-;_-@_-'
    PCT_FMT = '0.0"%"'

    def header_row(ws, row, headers):
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.fill = HEADER_FILL
            c.font = HEADER_FONT
            c.alignment = Alignment(horizontal="left", vertical="center")
            c.border = BORDER

    def fit_columns(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def write_kv(ws, row, key, value, money=False):
        ws.cell(row=row, column=1, value=key).font = SUB_FONT
        c = ws.cell(row=row, column=2, value=value)
        if money and isinstance(value, (int, float)):
            c.number_format = INR_FMT
        c.alignment = Alignment(horizontal="left")
        return row + 1

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Bank Statement Analysis"
    ws["A1"].font = Font(bold=True, size=16, color="0F172A")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y, %I:%M %p')}"
    ws["A2"].font = Font(italic=True, color="64748B", size=10)

    r = 4
    a = analysis["account"]
    t = analysis["totals"]
    bm = analysis["balance_metrics"]
    sc = analysis["score"]

    ws.cell(row=r, column=1, value="ACCOUNT & PERIOD").font = SECTION_FONT
    r += 1
    r = write_kv(ws, r, "Account holder", a.get("holder", "—"))
    r = write_kv(ws, r, "Filename", a.get("filename", "—"))
    r = write_kv(ws, r, "Period from", a["period"]["from"])
    r = write_kv(ws, r, "Period to", a["period"]["to"])
    r = write_kv(ws, r, "Months covered", a["period"]["months"])
    r = write_kv(ws, r, "Total transactions", a["transactions"])

    r += 1
    ws.cell(row=r, column=1, value="CREDITWORTHINESS SCORE").font = SECTION_FONT
    r += 1
    r = write_kv(ws, r, "Score (0-100)", sc["value"])
    r = write_kv(ws, r, "Rating", sc["rating"])
    r = write_kv(ws, r, "Scale", sc["scale"])

    r += 1
    ws.cell(row=r, column=1, value="CASH FLOW TOTALS").font = SECTION_FONT
    r += 1
    r = write_kv(ws, r, "Total Inflow (₹)", t.get("total_inflow"), money=True)
    r = write_kv(ws, r, "Total Outflow (₹)", t.get("total_outflow"), money=True)
    r = write_kv(ws, r, "Net Cash Flow (₹)", t.get("net_cash_flow"), money=True)
    r = write_kv(ws, r, "Credit transactions", t.get("credit_count"))
    r = write_kv(ws, r, "Debit transactions", t.get("debit_count"))
    r = write_kv(ws, r, "Opening Balance (₹)", t.get("opening_balance"), money=True)
    r = write_kv(ws, r, "Closing Balance (₹)", t.get("closing_balance"), money=True)

    r += 1
    ws.cell(row=r, column=1, value="BALANCE METRICS").font = SECTION_FONT
    r += 1
    r = write_kv(ws, r, "Average Balance (₹)", bm.get("average"), money=True)
    r = write_kv(ws, r, "Average Monthly Balance (AMB) (₹)", bm.get("amb"), money=True)
    r = write_kv(ws, r, "Minimum Balance (₹)", bm.get("minimum"), money=True)
    r = write_kv(ws, r, "Maximum Balance (₹)", bm.get("maximum"), money=True)
    r = write_kv(ws, r, "Volatility (%)", bm.get("volatility_pct"))
    r = write_kv(ws, r, "Days in negative balance", bm.get("negative_days"))

    fit_columns(ws, [38, 28])

    # ── Sheet 2: Risk Flags ──────────────────────────────────────────────────
    ws = wb.create_sheet("Risk Flags")
    header_row(ws, 1, ["Severity", "Code", "Title", "Detail"])
    SEV_COLOR = {"high": "FECACA", "medium": "FED7AA", "low": "BFDBFE"}
    for i, f in enumerate(analysis.get("risk_flags", []), start=2):
        ws.cell(row=i, column=1, value=f["severity"].upper())
        ws.cell(row=i, column=2, value=f["code"])
        ws.cell(row=i, column=3, value=f["title"])
        ws.cell(row=i, column=4, value=f["detail"])
        for col in range(1, 5):
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor=SEV_COLOR.get(f["severity"], "F1F5F9"))
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical="top")
            ws.cell(row=i, column=col).border = BORDER
    if not analysis.get("risk_flags"):
        ws.cell(row=2, column=1, value="No red flags detected.")
    fit_columns(ws, [12, 28, 50, 70])

    # ── Sheet 3: Underwriter Notes ───────────────────────────────────────────
    ws = wb.create_sheet("Underwriter Notes")
    header_row(ws, 1, ["#", "Observation"])
    for i, obs in enumerate(analysis.get("underwriter_observations", []), start=2):
        ws.cell(row=i, column=1, value=i - 1)
        c = ws.cell(row=i, column=2, value=obs)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDER
        ws.cell(row=i, column=1).border = BORDER
    fit_columns(ws, [6, 110])

    # ── Sheet 4: Monthly Trend ───────────────────────────────────────────────
    ws = wb.create_sheet("Monthly Trend")
    headers = ["Month", "Inflow (₹)", "Outflow (₹)", "Net (₹)", "Txns", "Closing Balance (₹)"]
    header_row(ws, 1, headers)
    for i, m in enumerate(analysis.get("monthly_trend", []), start=2):
        ws.cell(row=i, column=1, value=m["month"])
        ws.cell(row=i, column=2, value=m["inflow"]).number_format = INR_FMT
        ws.cell(row=i, column=3, value=m["outflow"]).number_format = INR_FMT
        ws.cell(row=i, column=4, value=m["net"]).number_format = INR_FMT
        ws.cell(row=i, column=5, value=m["txn_count"])
        ws.cell(row=i, column=6, value=m["closing_balance"]).number_format = INR_FMT
        for col in range(1, 7):
            ws.cell(row=i, column=col).border = BORDER
    fit_columns(ws, [12, 18, 18, 18, 8, 22])
    ws.freeze_panes = "A2"

    # ── Sheet 5: Transactions (full classified ledger) ───────────────────────
    ws = wb.create_sheet("Transactions")
    headers = ["Date", "Description", "Category", "Channel", "Counterparty", "Type",
               "Credit (₹)", "Debit (₹)", "Running Balance (₹)"]
    header_row(ws, 1, headers)
    for i, t in enumerate(analysis.get("transactions", []), start=2):
        ws.cell(row=i, column=1, value=t["date"])
        ws.cell(row=i, column=2, value=t["description"])
        ws.cell(row=i, column=3, value=t["category"])
        ws.cell(row=i, column=4, value=t["channel"])
        ws.cell(row=i, column=5, value=t["counterparty"])
        ws.cell(row=i, column=6, value=t["type"])
        c_cell = ws.cell(row=i, column=7, value=t["credit"])
        d_cell = ws.cell(row=i, column=8, value=t["debit"])
        b_cell = ws.cell(row=i, column=9, value=t["running_balance"])
        c_cell.number_format = INR_FMT
        d_cell.number_format = INR_FMT
        b_cell.number_format = INR_FMT
        if t["type"] == "Credit":
            c_cell.font = Font(color="047857")
        else:
            d_cell.font = Font(color="B91C1C")
        for col in range(1, 10):
            ws.cell(row=i, column=col).border = BORDER
            ws.cell(row=i, column=col).alignment = Alignment(vertical="top", wrap_text=(col == 2))
    fit_columns(ws, [12, 50, 26, 16, 28, 8, 16, 16, 20])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{ws.max_row}"

    # ── Sheet 6: Recurring Credits ───────────────────────────────────────────
    def _write_recurring_sheet(name, items, accent_color):
        s = wb.create_sheet(name)
        header_row(s, 1, ["Counterparty", "Median (₹)", "Total (₹)", "Occurrences", "First Seen", "Last Seen", "Sample Description"])
        for i, r in enumerate(items, start=2):
            s.cell(row=i, column=1, value=r["counterparty"])
            s.cell(row=i, column=2, value=r["median_amount"]).number_format = INR_FMT
            tot = s.cell(row=i, column=3, value=r["total_amount"])
            tot.number_format = INR_FMT
            tot.font = Font(bold=True, color=accent_color)
            s.cell(row=i, column=4, value=r["occurrences"])
            s.cell(row=i, column=5, value=r["first_seen"])
            s.cell(row=i, column=6, value=r["last_seen"])
            s.cell(row=i, column=7, value=r["sample_description"])
            for col in range(1, 8):
                s.cell(row=i, column=col).border = BORDER
        fit_columns(s, [32, 16, 16, 12, 14, 14, 60])
        s.freeze_panes = "A2"

    _write_recurring_sheet("Recurring Credits", analysis.get("recurring_credits", []), "047857")
    _write_recurring_sheet("Recurring Debits", analysis.get("recurring_debits", []), "B91C1C")
    _write_recurring_sheet("EMI Obligations", analysis.get("emi_obligations", []), "B45309")

    # ── Sheet 9: Cash Activity ───────────────────────────────────────────────
    ws = wb.create_sheet("Cash Activity")
    header_row(ws, 1, ["Type", "Count", "Total (₹)", "Largest Single (₹)"])
    cash = analysis.get("cash_activity", {})
    rows = [
        ("Cash Deposits", cash.get("cash_deposits", {})),
        ("Cash Withdrawals", cash.get("cash_withdrawals", {})),
    ]
    for i, (label, d) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=d.get("count", 0))
        ws.cell(row=i, column=3, value=d.get("total", 0)).number_format = INR_FMT
        ws.cell(row=i, column=4, value=d.get("max_single", 0)).number_format = INR_FMT
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = BORDER
    fit_columns(ws, [22, 10, 18, 22])

    # ── Sheet 10: Cheque Returns ─────────────────────────────────────────────
    ws = wb.create_sheet("Cheque Returns")
    header_row(ws, 1, ["Date", "Type", "Amount (₹)", "Description"])
    for i, b in enumerate(analysis.get("cheque_returns", {}).get("items", []), start=2):
        ws.cell(row=i, column=1, value=b["date"])
        ws.cell(row=i, column=2, value=b["type"])
        ws.cell(row=i, column=3, value=b["amount"]).number_format = INR_FMT
        ws.cell(row=i, column=4, value=b["description"])
        for col in range(1, 5):
            ws.cell(row=i, column=col).border = BORDER
            ws.cell(row=i, column=col).fill = PatternFill("solid", fgColor="FEE2E2")
    if not analysis.get("cheque_returns", {}).get("items"):
        ws.cell(row=2, column=1, value="No cheque/ECS returns detected.")
    fit_columns(ws, [12, 12, 16, 70])

    # ── Sheet 11: Counterparty Concentration ─────────────────────────────────
    ws = wb.create_sheet("Counterparty Concentration")
    conc = analysis.get("counterparty_concentration", {})
    ws.cell(row=1, column=1, value="TOP INFLOW COUNTERPARTIES").font = SECTION_FONT
    header_row(ws, 2, ["Counterparty", "Amount (₹)", "Share %"])
    r = 3
    for c in conc.get("top_inflow_counterparties", []):
        ws.cell(row=r, column=1, value=c["counterparty"])
        ws.cell(row=r, column=2, value=c["amount"]).number_format = INR_FMT
        ws.cell(row=r, column=3, value=c["share_pct"]).number_format = PCT_FMT
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BORDER
        r += 1
    r += 1
    ws.cell(row=r, column=1, value="TOP OUTFLOW COUNTERPARTIES").font = SECTION_FONT
    r += 1
    header_row(ws, r, ["Counterparty", "Amount (₹)", "Share %"])
    r += 1
    for c in conc.get("top_outflow_counterparties", []):
        ws.cell(row=r, column=1, value=c["counterparty"])
        ws.cell(row=r, column=2, value=c["amount"]).number_format = INR_FMT
        ws.cell(row=r, column=3, value=c["share_pct"]).number_format = PCT_FMT
        for col in range(1, 4):
            ws.cell(row=r, column=col).border = BORDER
        r += 1
    fit_columns(ws, [40, 18, 12])

    # Save
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
